import os
import fnmatch 
import pandas as pd
from openpyxl import load_workbook
import openpyxl

#######################################################
# Function to find out the RateDoc Files present in the Directory
#######################################################
def multiple_dfs(df_list, sheets, file_name, spaces):
    writer = pd.ExcelWriter(file_name,engine='xlsxwriter')   
    row = 0
    for dataframe in df_list:
        dataframe.to_excel(writer,sheet_name=sheets,startrow=row , startcol=0)   
        row = row + len(dataframe.index) + spaces + 1
    writer.save()


def file_finder():
    treeroot = r"D:\\"
    pattern = '*RateDoc.html'
    results = []                 # an emplty list 
    for base, dirs, files in os.walk(treeroot):
        goodfiles = fnmatch.filter(files, pattern)                       # fnmatch function to filter the files based on pattern assigned
        results.extend(os.path.join(base, f) for f in goodfiles)
    return results

file_found = file_finder()
# sheets = ['df1','df2','df3','df4','df5','df6','df7','df8'] 
for f in file_found:
    table_sheet = pd.read_html(f)
    df1 = table_sheet[0]
    df2 = table_sheet[1]
    dfs = [df1, df2]
    multiple_dfs(dfs, sheets[j], 'test1.xlsx', 1)
       
def dfs_tabs(df_list, sheet_list, file_name):
    writer = pd.ExcelWriter(file_name,engine='xlsxwriter')   
    for dataframe, sheet in zip(df_list, sheet_list):
        dataframe.to_excel(writer, sheet_name=sheet, startrow=0 , startcol=0)   
    writer.save()

# list of dataframes and sheet names
dfs = [df, df1, df2]
sheets = ['df','df1','df2']    

# run function
dfs_tabs(dfs, sheets, 'multi-test.xlsx')









# def opening_html():
#     file_found = file_finder()
#     # wb = openpyxl.Workbook()
#     # wb.save('data.xlsx')
#     for f in file_found:
#         # print(f)
#         table_sheet = pd.read_html(f)
#         print(type(table_sheet))
#         print(table_sheet)
#         # wb2 = load_workbook('data.xlsx')
#         # # wb2.create_sheet('sid1')
#         # table_sheet.to_excel(wb2,sheet_name="Sheet1")
#         # wb2.save('data.xlsx')
        
#         # table_sheet.to_excel("data.xlsx")
    

# opening_html()



# # def save_data_mastersheet(final):
# #     #final = match_unique()
# #     path = r"Book1.xlsx"
# #     book = load_workbook(path)
# #     writer = pd.ExcelWriter(path, engine='openpyxl')
# #     writer.book = book
# #     # if 'mastersheet' in book.sheetnames:
# #     #     pfd = book['mastersheet']
# #     #     book.remove(pfd)
# #     final.to_excel(writer, sheet_name='mastersheet')

# #     writer.save()
# #     writer.close()